# -*- encoding: utf-8 -*-
import urllib
import random
import webbrowser
from lxml import html
import pytagcloud
import pandas as pd
import simplejson
import sys
from konlpy.tag import Twitter
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium import common
from bs4 import BeautifulSoup
from collections import Counter
import time
import urllib
import re

"""
wb = Workbook()
sheet1 = wb.active
sheet1.title = '수집 데이터'
wb2 = Workbook()
sheet2 = wb2.active
sheet2.title = '분석 데이터'
"""
"""
r = lambda: random.randint(0,255)
# 글씨의 랜덤색깔
color = lambda: (r(), r(), r())

def get_tags_draw(text, ntags=50, multiplier=3):  # 전에 했던 명사 탐색
    spliter = Twitter()
    nouns = spliter.nouns(text)
    count = Counter(nouns)
    return [{'color': color(), 'tag': n, 'size': (c * multiplier) / 2} \
            for n, c in count.most_common(ntags)]

def draw_cloud(tags, filename, fontname='NanumGothic', size=(800, 600)):
    pytagcloud.create_tag_image(tags, filename, fontname=fontname, size=size)
    webbrowser.open(filename)
"""

def get_tags(text, ntags=100):
    # konlpy의 Twitter객체
    spliter = Twitter()
    # nouns 함수를 통해서 text에서 명사만 분리/추출
    nouns = spliter.nouns(text)
    # Counter객체를 생성하고 참조변수 nouns할당
    count = Counter(nouns)
    # 명사 빈도수 저장할 변수
    return_list = []
    for n, c in count.most_common(ntags):
        temp = {'tag': n, 'count': c}
        return_list.append(temp)
    return return_list


def clean_text(text):
    change_text= ['고혈압','음식','것','등','수','건강','경우','이','섭취','식품','병원','때문'
                  ,'질환','예방','및','치료','때','환자','효과','사람','심장','그','중','도움','약'
                  ,'성분','운동','더','의','몸','후','증상','정도','교수','결과','원인','말','발생','방법'
                  ,'이상','가장','은','안','습관','기능','위험','관리','데','명','연구','또한','세포'
                  ,'함유','우리','효능','수치','물질','비만','증가','일','통해','도','위해','수술','작용'
                  ,'팀','미국','사용','치매','자신','복용','방송','제거','하루','대해','과','영향','문제'
                  ,'여성','상태','정보','알','생각','시간','유지','속','사실','솔잎','개선','고','법','서울'
                  ,'뉴스','대표','내','박사','생활','식사','유발','기','피','모두','대한','다른','식단','조절'
                  ,'생','상승','확인','를','최근','주의','날','제','번','심','포함','로','식'
                  ,'상','정','려','전','리','교','해','자','사','나','주','위','인','선','체','성'
                  ,'적','줄','시','증','치','부','장','용','관','공개','국','런','러','름','보','외','억'
                  ,'초','개','유','요','료','양','요한','세','거','두','분','산','병','살','만','뿐'
                  ,'준','오','또','쥐','소개','대부분','역할','정상','하나','민','진행','품','신체','발표'
                  ,'코','국민연금','관련','거나','감소','활동','한국','류','칼로리','량','분석','자주','점'
                  ,'따라서','정신','반','불','구','부분','다음','항','대상','개월','대학','남성','촉진','집'
                  ,'지난','시작','원','매우','축','섬유','요리','실','가슴','계','진단']
    cleaned_text = re.sub('[a-zA-Z]', '', text)
    cleaned_text = re.sub('[\{\}\[\]\/?.,;:|\)*~`!^\-_+<>@\#$%&\\\=\(\'\"]',
                          '', cleaned_text)
    for i in range(len(change_text)):
        cleaned_text = re.sub(change_text[i], '', cleaned_text)
    return cleaned_text

def main():

    text_file_name = "피부탄력에 피해야할 음식.txt"
    # 최대 많은 빈도수 부터 20개 명사 추출
    noun_count = 400
    # count.txt 에 저장
    output_file_name = "피부탄력에 피해야할 음식_count.txt"
    open_text_file = open(text_file_name, 'r', -1, encoding='utf-8')
    text = open_text_file.read()
    text=clean_text(text)

    tags = get_tags(text, noun_count)  # get_tags 함수 실행

    open_text_file.close()
    #ws.close()
    #rb.close()# 파일 close
    # 결과로 쓰일 count.txt 열기
    open_output_file = open(output_file_name, 'w', -1, encoding='utf-8')
    for tag in tags:
        noun = tag['tag']
        count = tag['count']
        #sheet2.append([tag.text])
        open_output_file.write('{} {}\n'.format(noun, count))
        # 결과 저장
    #wb2.save('info.xlsx')
    open_output_file.close()

if __name__ == '__main__':
    main()