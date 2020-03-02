# -*- encoding: utf-8 -*-
from konlpy.tag import Twitter
from openpyxl import load_workbook
from selenium import webdriver
from selenium import common
from bs4 import BeautifulSoup
from collections import Counter
import time
import openpyxl
#메모장 저장



f = open("피부탄력에 피해야할 음식.txt",'w',-1,"utf-8")

#엑셀 저장용 변수
wb = openpyxl.Workbook()
sheet1 = wb.active
sheet1.title = '수집 데이터'
#wb2 = openpyxl.Workbook()
#sheet1 = wb2.active
#sheet1.title = '분석 데이터'
#web창 2개 생성
driver = webdriver.Chrome('./chromedriver')
driver2 = webdriver.Chrome('./chromedriver')
#web창 1개는 구글검색용
driver.get('https://google.com/')
#검색창에 입력
driver.find_element_by_css_selector('div.a4bIc > input').send_keys('피부탄력에 피해야할 음식')
time.sleep(2)
#검새버튼 클릭
#driver.find_element_by_css_selector('div.gb_ld > a').click()
driver.find_element_by_xpath('//*[@id="tsf"]/div[2]/div[1]/div[3]/center/input[1]').click()
#driver.find_element_by_name('btnK').click


time.sleep(1)

#페이지 넘기기 변수
page=1
"""
def get_tags(text, ntags=50):
    # konlpy의 Twitter객체
    spliter = Twitter()
    # nouns 함수를 통해서 text에서 명사만 분리/추출
    nouns = spliter.nouns(text)
    # Counter객체를 생성하고 참조변수 nouns할당
    count = Counter(nouns)
    # 명사 빈도수 저장할 변수
    return_list = []
    for n, c in count.most_common(ntags):
        temp = {'tag': n, 'count': c}
        return_list.append(temp)
    return return_list

def main():
    # 분석할 파일
    rb=load_workbook("D:\jooho\python_64\py_project\test\test3.xlsx")
    ws=rb['수집 데이터']
    # 최대 많은 빈도수 부터 20개 명사 추출
    noun_count = 20
    # count.txt 에 저장
    output_file_name = "count.txt"
    # 분석할 파일을 open
    ws=rb['수집 데이터']
    text =ws.value  # 파일을 읽습니다.
    tags = get_tags(text, noun_count)  # get_tags 함수 실행
    ws.close()
    rb.close()# 파일 close
    # 결과로 쓰일 count.txt 열기
    #open_output_file = open(output_file_name, 'r', -1, "utf-8")
    for tag in tags:
        noun = tag['tag']
        count = tag['count']
        sheet1.append([tag.text])
       # open_output_file.write('{} {}\n'.format(noun, count))
        # 결과 저장
    wb2.save('info.xlsx')
    #open_output_file.close()
"""
while True:
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    #검색 전체목록
    #list = driver.find_elements_by_css_selector('div.srg > div.g')
    #검색 전체목록
    list2 = soup.select('div.srg > div.g')
    #페이지 들어가서 내용
    list3 = soup.select('p')
    for data in list2:
        #url을 가져오기
        link = data.select_one('div.rc > div.r > a').attrs['href']
        #가져온 url 실행
        driver2.get(link)
        time.sleep(1)
        #for data2 in list3:
            #페이지 내용을 texts에 담기
        results=driver2.find_elements_by_css_selector('p')
            #results=data2.text
        for t in results:
            print(t.text)
            sheet1.append([t.text])
            f.write(t.text)



    page = page + 1
    try:
        if(page==8):
            break
        else:
            driver.find_element_by_xpath('//a[text()=' + str(page) + ']').click()

    except common.exceptions.NoSuchElementException:
        wb.save('test3.xlsx')
        break


#data.find_element_by_css_selector('div.rc > div.r > a').click()
#driver.find_element_by_xpath("//a[@class='link_navbar search']").click()
#driver.find_element_by_xpath("//button[@class='btn_search ng-star-inserted']").click()
